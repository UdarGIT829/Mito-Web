import argparse
import json
import re
import sqlite3
import sys
import threading
import time
from collections.abc import Iterator
from dataclasses import dataclass, field
from pathlib import Path

import vcf_parser
import annotation_api


MITOCHONDRIAL_LENGTH = 16569
DEFAULT_MITO_REFERENCE_PATH = (
    Path(__file__).resolve().parent / "reference" / "hg38_chrM.fa"
)
DEFAULT_SUBJECT_REGEX = r"^(.*)$"
DEFAULT_ANNOTATION_DATABASE_PATH = (
    Path(__file__).resolve().parent / "mutation_annotations.sqlite"
)
ANNOTATION_PROVIDERS = ("ensembl", "clinvar", "mitomap")
SKIPPED_ANNOTATION_PROVIDERS = {
    "mitomap": "Skipping until api works",
}


class mitochondrial_base_positions(Iterator):
    def __init__(self):
        self._positions = iter(range(1, MITOCHONDRIAL_LENGTH + 1))

    def __iter__(self):
        return self

    def __next__(self):
        return next(self._positions)


@dataclass(frozen=True)
class MutationAllele:
    """Comparable mutation allele used for sample set operations."""

    position: int
    alt: str
    ref: str = field(compare=False)
    af: float | str | None = field(default=None, compare=False)
    filter: str = field(default="", compare=False)
    mutation: vcf_parser.VCFMutation | None = field(
        default=None,
        compare=False,
        repr=False,
    )


@dataclass
class Sample:
    sample_id: str
    population: list[str]
    source_path: Path
    mutations: list[vcf_parser.VCFMutation] = field(default_factory=list)

    def __contains__(self, tags):
        """Return True when this sample has the given population tag(s)."""
        if isinstance(tags, str):
            return tags in self.population

        return all(tag in self.population for tag in tags)

    def __iter__(self):
        """Iterate over population tags."""
        return iter(self.population)

    @property
    def population_key(self):
        """Return a stable serialized population tag key."""
        return "|".join(self.population)

    @property
    def label(self):
        """Return a compact subject/population label."""
        if not self.population:
            return self.sample_id
        return f"{self.sample_id}_{'_'.join(self.population)}"

    def has_any(self, tags):
        """Return True when this sample has at least one of the given tags."""
        return any(tag in self for tag in tags)

    def has_all(self, tags):
        """Return True when this sample has every given tag."""
        return tags in self

    def is_subject(self, sample_id):
        """Return True when this sample belongs to the given subject."""
        return self.sample_id == sample_id

    @property
    def mutation_alleles(self):
        """Return comparable mutation alleles for this sample."""
        alleles = set()
        for mutation in self.mutations:
            for index, alt in enumerate(mutation.alts):
                af = mutation.afs[index] if index < len(mutation.afs) else None
                alleles.add(MutationAllele(
                    position=mutation.position,
                    ref=mutation.ref,
                    alt=alt,
                    af=af,
                    filter=mutation.filter,
                    mutation=mutation,
                ))
        return alleles

    def __and__(self, other):
        """Return mutation alleles common to both samples."""
        return self.mutation_alleles & other.mutation_alleles

    def __or__(self, other):
        """Return mutation alleles present in either sample."""
        return self.mutation_alleles | other.mutation_alleles

    def __sub__(self, other):
        """Return mutation alleles present in this sample but not the other."""
        return self.mutation_alleles - other.mutation_alleles

    def __xor__(self, other):
        """Return mutation alleles that differ between two samples."""
        return self.mutation_alleles ^ other.mutation_alleles

    def common_mutations(self, *others):
        """Return mutation alleles common to this sample and all others."""
        common = self.mutation_alleles
        for other in others:
            common &= other.mutation_alleles
        return common

    def different_mutations(self, *others):
        """Return mutation alleles not shared by every provided sample."""
        samples = (self, *others)
        if not samples:
            return set()

        union = set()
        common = samples[0].mutation_alleles
        for sample in samples:
            union |= sample.mutation_alleles
            common &= sample.mutation_alleles

        return union - common


@dataclass
class ImportPlanItem:
    """One VCF file and the metadata that will be used during import."""

    source_path: Path
    sample_id: str
    population: list[str] = field(default_factory=list)


def is_vcf_file(path):
    """Return True for plain or gzipped VCF files."""
    name = path.name
    return name.endswith(".vcf") or name.endswith(".vcf.gz")


def strip_vcf_suffix(path):
    """Return a VCF filename without .vcf or .vcf.gz."""
    name = Path(path).name
    if name.endswith(".vcf.gz"):
        return name[:-len(".vcf.gz")]
    if name.endswith(".vcf"):
        return name[:-len(".vcf")]
    return Path(path).stem


def default_sample_id(path):
    """Return a generic sample ID inferred from a VCF filename."""
    stem = strip_vcf_suffix(path)
    return re.sub(r"_S\d+$", "", stem)


def parse_sample_filename(path, input_dir=None):
    """Return generic sample metadata inferred from a VCF path."""
    sample_id = default_sample_id(path)
    population = []
    if input_dir is not None:
        try:
            relative_parent = Path(path).parent.relative_to(Path(input_dir))
        except ValueError:
            relative_parent = Path()
        population = [
            part
            for part in relative_parent.parts
            if part and part != "."
        ]

    return sample_id, population


def is_dna_motif(motif):
    """Return True when all motif characters are DNA bases."""
    return bool(motif) and all(base in {"A", "C", "G", "T", "N"} for base in motif)


def is_minimal_motif(motif):
    """Return True when a motif is not made from a smaller repeated motif."""
    for size in range(1, len(motif)):
        if len(motif) % size:
            continue
        smaller = motif[:size]
        if smaller * (len(motif) // size) == motif:
            return False
    return True


def longest_repeat(sequence, motif_length):
    """Return the longest repeated motif run for a fixed motif length."""
    sequence = sequence.upper()
    best_motif = ""
    best_count = 0
    limit = len(sequence) - motif_length + 1

    for start in range(max(limit, 0)):
        motif = sequence[start:start + motif_length]
        if not is_dna_motif(motif) or not is_minimal_motif(motif):
            continue

        count = 1
        next_start = start + motif_length
        while sequence[next_start:next_start + motif_length] == motif:
            count += 1
            next_start += motif_length

        if count > best_count and count > 1:
            best_motif = motif
            best_count = count

    return best_motif, best_count


def repeat_profile(sequence):
    """Return 1-, 2-, and 3-base repeat motifs and counts for a sequence."""
    return {
        motif_length: longest_repeat(sequence, motif_length)
        for motif_length in (1, 2, 3)
    }


def strongest_alt_repeat_profile(alts):
    """Return the strongest 1-, 2-, and 3-base repeats across ALT alleles."""
    strongest = {
        1: ("", 0),
        2: ("", 0),
        3: ("", 0),
    }

    for alt in alts:
        for motif_length, (motif, count) in repeat_profile(alt).items():
            if count > strongest[motif_length][1]:
                strongest[motif_length] = (motif, count)

    return strongest


def reference_context(reference, position, flank_size=6):
    """Return reference bases before and after a 1-indexed circular position."""
    reference_length = len(reference)

    def base_at(wrapped_position):
        index = (wrapped_position - 1) % reference_length + 1
        return reference[index]

    before = "".join(
        base_at(position - offset)
        for offset in range(flank_size, 0, -1)
    )
    after = "".join(
        base_at(position + offset)
        for offset in range(1, flank_size + 1)
    )

    return before, after


def iter_variant_files(input_dir):
    """Yield one VCF per sample from an input directory."""
    base_dir = Path(input_dir)
    if not base_dir.exists():
        raise FileNotFoundError(f"Input directory not found: {base_dir}")

    selected_paths = {}
    for path in sorted(base_dir.rglob("*")):
        if not path.is_file() or not is_vcf_file(path):
            continue

        sample_id, population = parse_sample_filename(path, base_dir)
        sample_key = (sample_id, tuple(population))
        existing_path = selected_paths.get(sample_key)
        if existing_path is None or existing_path.suffix == ".gz":
            selected_paths[sample_key] = path

    for path in selected_paths.values():
        yield path


def build_import_plan(input_dir):
    """Return editable import metadata for all logical VCF samples."""
    plan = []
    for path in iter_variant_files(input_dir):
        sample_id, population = parse_sample_filename(path, input_dir)
        plan.append(ImportPlanItem(
            source_path=path,
            sample_id=sample_id,
            population=population,
        ))
    return plan


def split_tags(text):
    """Return normalized tags from comma, semicolon, or pipe separated text."""
    return [
        tag.strip()
        for tag in re.split(r"[,;|]", text)
        if tag.strip()
    ]


def subject_id_from_regex(path, pattern):
    """Return a subject ID extracted from a VCF filename with a regex."""
    text = strip_vcf_suffix(path)
    match = re.search(pattern, text)
    if not match:
        return ""

    if match.groups():
        return (match.group(1) or "").strip()

    return (match.group(0) or "").strip()


def load_sample(path, input_dir=None):
    """Build a Sample from a filtered VCF path and its mutations."""
    sample_id, population = parse_sample_filename(path, input_dir)
    sample = Sample(
        sample_id=sample_id,
        population=population,
        source_path=Path(path),
    )

    with vcf_parser.VCFIterator(str(path)) as rows:
        for mutation in rows:
            sample.mutations.append(mutation)

    return sample


def load_sample_from_plan(item):
    """Build a Sample from editable import metadata."""
    sample = Sample(
        sample_id=item.sample_id,
        population=list(item.population),
        source_path=Path(item.source_path),
    )

    with vcf_parser.VCFIterator(str(item.source_path)) as rows:
        for mutation in rows:
            sample.mutations.append(mutation)

    return sample


def load_samples(input_dir):
    """Load all logical samples from a variant call set."""
    return [
        load_sample(path, input_dir)
        for path in iter_variant_files(input_dir)
    ]


def load_samples_from_plan(plan):
    """Load all samples from an editable import plan."""
    return [load_sample_from_plan(item) for item in plan]


def load_mito_reference(path=DEFAULT_MITO_REFERENCE_PATH):
    """Load a mitochondrial FASTA sequence as a 1-indexed position map."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(
            f"Missing mitochondrial reference FASTA: {path}. "
            "Run scripts/download_mito_ref.py first."
        )

    sequence_parts = []
    with path.open() as handle:
        for line in handle:
            line = line.strip()
            if not line or line.startswith(">"):
                continue
            sequence_parts.append(line.upper())

    sequence = "".join(sequence_parts)
    if len(sequence) != MITOCHONDRIAL_LENGTH:
        raise ValueError(
            f"Expected mitochondrial reference length {MITOCHONDRIAL_LENGTH}, "
            f"got {len(sequence)} from {path}"
        )

    return {
        position: base
        for position, base in enumerate(sequence, start=1)
    }


def summarize_sample(sample):
    """Return basic mutation counts for a loaded sample."""
    total_variants = len(sample.mutations)
    pass_variants = sum(
        1 for mutation in sample.mutations
        if mutation.filter == "PASS"
    )
    multi_alt_variants = sum(
        1 for mutation in sample.mutations
        if mutation.has_multiple_alt_alleles
    )
    example_metadata = sample.mutations[-1].metadata if sample.mutations else {}

    return total_variants, pass_variants, multi_alt_variants, example_metadata


def format_mutation(mutation):
    """Return one readable line for a parsed VCF mutation."""
    return (
        f"  POS {mutation.position}: "
        f"{mutation.ref}>{mutation.alt} "
        f"FILTER={mutation.filter} "
        f"ALTS={mutation.alts} "
        f"AFS={mutation.afs} "
        f"ALT_AFS={mutation.alt_afs} "
        f"METADATA={mutation.metadata}"
    )


def sample_label(sample):
    """Return a compact label for sample-specific output columns."""
    return sample.label


def append_unique(values, value):
    """Append value if present and not already in values."""
    if value and value not in values:
        values.append(value)


def mutation_alt_text(mutations):
    """Return comma-packed ALT values for one sample at one position."""
    return ",".join(
        mutation.alt
        for mutation in mutations
        if mutation.alt
    )


def mutation_af_text(mutations):
    """Return comma-packed AF values for one sample at one position."""
    return ",".join(
        mutation.allele_fraction_text
        for mutation in mutations
        if mutation.allele_fraction_text
    )


def mutation_metadata_text(mutation):
    """Return mutation metadata as semicolon-delimited key=value pairs."""
    return ";".join(
        f"{key}={value}"
        for key, value in mutation.metadata.items()
    )


def mutation_alt_af_text(mutation):
    """Return ALT/AF pairs as semicolon-delimited key=value pairs."""
    return ";".join(
        f"{alt}={af}"
        for alt, af in mutation.alt_afs.items()
    )


def safe_sheet_title(title, existing_titles):
    """Return an Excel-safe sheet title, unique within a workbook."""
    invalid_chars = "[]:*?/\\"
    cleaned = "".join("_" if char in invalid_chars else char for char in title)
    cleaned = cleaned[:31] or "Sheet"

    if cleaned not in existing_titles:
        return cleaned

    counter = 2
    while True:
        suffix = f"_{counter}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        if candidate not in existing_titles:
            return candidate
        counter += 1


def append_sample_sheet(workbook, sample, reference):
    """Append one worksheet containing all mutations from one VCF sample."""
    title = safe_sheet_title(sample_label(sample), set(workbook.sheetnames))
    worksheet = workbook.create_sheet(title)
    worksheet.append([
        "pos",
        "ref",
        "vcf ref",
        "alt",
        "af",
        "filter",
        "alts",
        "afs",
        "alt_afs",
        "metadata",
        "source file",
    ])

    for mutation in sample.mutations:
        worksheet.append([
            mutation.position,
            reference.get(mutation.position, ""),
            mutation.ref,
            mutation.alt,
            mutation.allele_fraction_text,
            mutation.filter,
            ",".join(mutation.alts),
            ",".join(str(af) for af in mutation.afs),
            mutation_alt_af_text(mutation),
            mutation_metadata_text(mutation),
            sample.source_path.name,
        ])


def make_master_xlsx(
    output_path,
    input_dir,
    reference_path=DEFAULT_MITO_REFERENCE_PATH,
):
    """Write a master mutation table as an XLSX workbook."""
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required to write master.xlsx. "
            "Install it with: python3 -m pip install openpyxl"
        ) from exc

    samples = load_samples(input_dir)
    reference = load_mito_reference(reference_path)
    mutation_lookup = {}

    for sample in samples:
        label = sample_label(sample)
        for mutation in sample.mutations:
            mutation_lookup.setdefault((label, mutation.position), []).append(mutation)

    headers = ["pos", "ref"]
    for sample in samples:
        label = sample_label(sample)
        headers.extend([f"alt of {label}", f"af of {label}"])

    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet("Master")
    worksheet.append(headers)

    for position in mitochondrial_base_positions():
        row = [position, reference[position]]
        for sample in samples:
            label = sample_label(sample)
            mutations = mutation_lookup.get((label, position), [])
            row.extend([
                mutation_alt_text(mutations),
                mutation_af_text(mutations),
            ])
        worksheet.append(row)

    for sample in samples:
        append_sample_sheet(workbook, sample, reference)

    output_path = Path(output_path)
    workbook.save(output_path)
    return output_path


def create_database_schema(connection):
    """Create the subject/sample/mutation database schema."""
    connection.executescript("""
        PRAGMA foreign_keys = ON;

        DROP TABLE IF EXISTS mutation_alts;
        DROP TABLE IF EXISTS mutations;
        DROP TABLE IF EXISTS sample_population_tags;
        DROP TABLE IF EXISTS samples;
        DROP TABLE IF EXISTS subjects;

        CREATE TABLE subjects (
            id INTEGER PRIMARY KEY,
            subject_id TEXT NOT NULL UNIQUE
        );

        CREATE TABLE samples (
            id INTEGER PRIMARY KEY,
            subject_id INTEGER NOT NULL,
            population_key TEXT NOT NULL,
            source_file TEXT NOT NULL,
            UNIQUE(subject_id, population_key),
            FOREIGN KEY(subject_id) REFERENCES subjects(id)
        );

        CREATE TABLE sample_population_tags (
            sample_id INTEGER NOT NULL,
            tag TEXT NOT NULL,
            tag_order INTEGER NOT NULL,
            PRIMARY KEY(sample_id, tag),
            FOREIGN KEY(sample_id) REFERENCES samples(id)
        );

        CREATE TABLE mutations (
            id INTEGER PRIMARY KEY,
            sample_id INTEGER NOT NULL,
            pos INTEGER NOT NULL,
            ref TEXT NOT NULL,
            vcf_ref TEXT NOT NULL,
            alt TEXT NOT NULL,
            af TEXT NOT NULL,
            polymorphism INTEGER NOT NULL,
            repeat_base TEXT NOT NULL,
            repeat_count INTEGER NOT NULL,
            repeat_2_bases TEXT NOT NULL,
            repeat_2_count INTEGER NOT NULL,
            repeat_3_bases TEXT NOT NULL,
            repeat_3_count INTEGER NOT NULL,
            filter TEXT NOT NULL,
            metadata_json TEXT NOT NULL,
            FOREIGN KEY(sample_id) REFERENCES samples(id)
        );

        CREATE TABLE mutation_alts (
            mutation_id INTEGER NOT NULL,
            alt_index INTEGER NOT NULL,
            alt TEXT NOT NULL,
            af REAL,
            af_text TEXT,
            polymorphism INTEGER NOT NULL,
            repeat_base TEXT NOT NULL,
            repeat_count INTEGER NOT NULL,
            repeat_2_bases TEXT NOT NULL,
            repeat_2_count INTEGER NOT NULL,
            repeat_3_bases TEXT NOT NULL,
            repeat_3_count INTEGER NOT NULL,
            PRIMARY KEY(mutation_id, alt_index),
            FOREIGN KEY(mutation_id) REFERENCES mutations(id)
        );

        CREATE INDEX idx_samples_subject_id ON samples(subject_id);
        CREATE INDEX idx_sample_population_tags_tag ON sample_population_tags(tag);
        CREATE INDEX idx_mutations_sample_pos ON mutations(sample_id, pos);
        CREATE INDEX idx_mutations_pos ON mutations(pos);
        CREATE INDEX idx_mutation_alts_alt ON mutation_alts(alt);
    """)


def create_annotation_database_schema(connection):
    """Create or upgrade the persistent, cross-import annotation cache."""
    connection.executescript("""
        PRAGMA foreign_keys = ON;

        CREATE TABLE IF NOT EXISTS annotation_variants (
            id INTEGER PRIMARY KEY,
            pos INTEGER NOT NULL,
            ref TEXT NOT NULL,
            alt TEXT NOT NULL,
            first_seen_at TEXT NOT NULL,
            last_seen_at TEXT NOT NULL,
            UNIQUE(pos, ref, alt)
        );

        CREATE TABLE IF NOT EXISTS provider_annotations (
            variant_id INTEGER NOT NULL,
            provider TEXT NOT NULL,
            annotation_json TEXT,
            error TEXT,
            retrieved_at TEXT,
            last_attempted_at TEXT NOT NULL,
            PRIMARY KEY(variant_id, provider),
            FOREIGN KEY(variant_id) REFERENCES annotation_variants(id)
        );

        CREATE INDEX IF NOT EXISTS idx_annotation_variants_pos_alt
        ON annotation_variants(pos, alt);
    """)


def imported_annotation_alleles(samples):
    """Return every unique VCF allele represented by imported mutations."""
    return sorted({
        (mutation.position, mutation.ref.upper(), alt.upper())
        for sample in samples
        for mutation in sample.mutations
        for alt in mutation.alts
        if alt and alt != "."
    })


def annotate_imported_mutations(
    annotation_db_path,
    samples,
    *,
    fetcher=None,
):
    """Fetch/cache annotations for alleles in newly imported samples."""
    return annotate_alleles(
        annotation_db_path,
        imported_annotation_alleles(samples),
        fetcher=fetcher,
    )


def annotate_alleles(
    annotation_db_path,
    alleles,
    *,
    fetcher=None,
    progress_callback=None,
):
    """Fetch missing annotations into one cache shared by all imports.

    Successful provider responses are never fetched again. Failed or missing
    providers remain eligible for retry on the next importer execution.
    """
    annotation_db_path = Path(annotation_db_path)
    annotation_db_path.parent.mkdir(parents=True, exist_ok=True)
    alleles = sorted(set(alleles))
    total_steps = len(alleles) * len(ANNOTATION_PROVIDERS)
    completed_steps = 0

    def report(position, ref, alt, provider, status):
        nonlocal completed_steps
        completed_steps += 1
        if progress_callback is not None:
            progress_callback(
                completed_steps,
                total_steps,
                position,
                ref,
                alt,
                provider,
                status,
            )

    with sqlite3.connect(annotation_db_path) as connection:
        create_annotation_database_schema(connection)
        for position, ref, alt in alleles:
            now = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
            connection.execute(
                """
                INSERT INTO annotation_variants(
                    pos, ref, alt, first_seen_at, last_seen_at
                ) VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(pos, ref, alt) DO UPDATE SET
                    last_seen_at = excluded.last_seen_at
                """,
                (position, ref, alt, now, now),
            )
            variant_id = connection.execute(
                "SELECT id FROM annotation_variants WHERE pos = ? AND ref = ? AND alt = ?",
                (position, ref, alt),
            ).fetchone()[0]
            completed = {
                row[0]
                for row in connection.execute(
                    """
                    SELECT provider FROM provider_annotations
                    WHERE variant_id = ? AND annotation_json IS NOT NULL
                      AND error IS NULL
                    """,
                    (variant_id,),
                )
            }
            missing = set(ANNOTATION_PROVIDERS) - completed
            for provider in ANNOTATION_PROVIDERS:
                if provider in completed:
                    report(position, ref, alt, provider, "cached")
            if not missing:
                continue

            results = {
                provider: {
                    "source": provider,
                    "status": message,
                    "skipped": True,
                }
                for provider, message in SKIPPED_ANNOTATION_PROVIDERS.items()
                if provider in missing
            }
            fetch_missing = missing - set(SKIPPED_ANNOTATION_PROVIDERS)

            if fetcher is not None and fetch_missing:
                fetched_results = fetcher(
                    position, ref, alt, continue_on_error=True
                )
                results.update({
                    provider: fetched_results.get(provider)
                    for provider in fetch_missing
                })
            else:
                provider_fetchers = {
                    "ensembl": lambda: annotation_api.fetch_ensembl_vep(
                        "MT", position, ref, alt
                    ),
                    "clinvar": lambda: annotation_api.fetch_clinvar_mito_variant(
                        position, ref, alt
                    ),
                    "mitomap": lambda: annotation_api.fetch_mitomap(
                        position, ref, alt
                    ),
                }
                for provider in fetch_missing:
                    try:
                        results[provider] = provider_fetchers[provider]()
                    except (annotation_api.AnnotationAPIError, ValueError) as exc:
                        results[provider] = {"source": provider, "error": str(exc)}

            attempted_at = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
            for provider in missing:
                result = results.get(provider)
                error = None
                annotation_json = None
                retrieved_at = None
                if result is None:
                    error = "Provider returned no result"
                elif isinstance(result, dict) and result.get("error"):
                    error = str(result["error"])
                else:
                    annotation_json = json.dumps(result, sort_keys=True)
                    if isinstance(result, dict):
                        retrieved_at = result.get("retrieved_at")
                connection.execute(
                    """
                    INSERT INTO provider_annotations(
                        variant_id, provider, annotation_json, error,
                        retrieved_at, last_attempted_at
                    ) VALUES (?, ?, ?, ?, ?, ?)
                    ON CONFLICT(variant_id, provider) DO UPDATE SET
                        annotation_json = excluded.annotation_json,
                        error = excluded.error,
                        retrieved_at = excluded.retrieved_at,
                        last_attempted_at = excluded.last_attempted_at
                    """,
                    (
                        variant_id, provider, annotation_json, error,
                        retrieved_at, attempted_at,
                    ),
                )
                connection.commit()
                report(
                    position,
                    ref,
                    alt,
                    provider,
                    (
                        "error"
                        if error
                        else "skipped"
                        if provider in SKIPPED_ANNOTATION_PROVIDERS
                        else "saved"
                    ),
                )
    return annotation_db_path


def annotation_alleles_from_database(database_path):
    """Read unique alleles from a current or legacy imported study database."""
    database_path = Path(database_path)
    if not database_path.is_file():
        raise FileNotFoundError(f"Database not found: {database_path}")

    with sqlite3.connect(database_path) as connection:
        tables = {
            row[0]
            for row in connection.execute(
                "SELECT name FROM sqlite_master WHERE type = 'table'"
            )
        }
        if "mutations" not in tables:
            raise ValueError(f"{database_path.name} has no mutations table")

        mutation_columns = {
            row[1] for row in connection.execute("PRAGMA table_info(mutations)")
        }
        if "pos" not in mutation_columns:
            raise ValueError(f"{database_path.name} has no mutations.pos column")
        ref_column = "vcf_ref" if "vcf_ref" in mutation_columns else "ref"
        if ref_column not in mutation_columns:
            raise ValueError(f"{database_path.name} has no mutation reference column")

        if "mutation_alts" in tables:
            alt_columns = {
                row[1]
                for row in connection.execute("PRAGMA table_info(mutation_alts)")
            }
        else:
            alt_columns = set()

        if {"mutation_id", "alt"} <= alt_columns:
            rows = connection.execute(
                f"""
                SELECT mutations.pos, mutations.{ref_column}, mutation_alts.alt
                FROM mutation_alts
                JOIN mutations ON mutations.id = mutation_alts.mutation_id
                """
            )
        elif "alt" in mutation_columns:
            rows = connection.execute(
                f"SELECT pos, {ref_column}, alt FROM mutations"
            )
        else:
            raise ValueError(f"{database_path.name} has no mutation ALT column")

        alleles = set()
        for position, ref, alt_text in rows:
            for alt in str(alt_text or "").split(","):
                alt = alt.strip().upper()
                if alt and alt != ".":
                    alleles.add((int(position), str(ref or "").upper(), alt))
        return alleles


def annotate_existing_databases(
    annotation_db_path,
    database_paths,
    *,
    fetcher=None,
    progress_callback=None,
):
    """Populate the shared cache from one or more previously imported DBs."""
    annotation_db_path = Path(annotation_db_path)
    alleles = set()
    for database_path in database_paths:
        database_path = Path(database_path)
        if database_path.resolve() == annotation_db_path.resolve():
            raise ValueError("The annotation database cannot be used as a study database.")
        alleles.update(annotation_alleles_from_database(database_path))
    annotate_alleles(
        annotation_db_path,
        alleles,
        fetcher=fetcher,
        progress_callback=progress_callback,
    )
    return annotation_db_path, len(alleles)


def clear_annotation_database(annotation_db_path):
    """Remove all cached variants and provider responses without deleting the DB."""
    annotation_db_path = Path(annotation_db_path)
    annotation_db_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(annotation_db_path) as connection:
        create_annotation_database_schema(connection)
        provider_count = connection.execute(
            "SELECT COUNT(*) FROM provider_annotations"
        ).fetchone()[0]
        variant_count = connection.execute(
            "SELECT COUNT(*) FROM annotation_variants"
        ).fetchone()[0]
        connection.execute("DELETE FROM provider_annotations")
        connection.execute("DELETE FROM annotation_variants")
    return variant_count, provider_count


def insert_subject(connection, subject_id):
    """Insert or fetch a subject row."""
    connection.execute(
        "INSERT OR IGNORE INTO subjects(subject_id) VALUES (?)",
        (subject_id,),
    )
    cursor = connection.execute(
        "SELECT id FROM subjects WHERE subject_id = ?",
        (subject_id,),
    )
    return cursor.fetchone()[0]


def insert_sample(connection, sample, subject_db_id):
    """Insert or fetch a sample row for a subject/population combination."""
    connection.execute(
        """
        INSERT OR IGNORE INTO samples(subject_id, population_key, source_file)
        VALUES (?, ?, ?)
        """,
        (subject_db_id, sample.population_key, sample.source_path.name),
    )
    cursor = connection.execute(
        """
        SELECT id FROM samples
        WHERE subject_id = ? AND population_key = ?
        """,
        (subject_db_id, sample.population_key),
    )
    sample_db_id = cursor.fetchone()[0]

    for tag_order, tag in enumerate(sample.population):
        connection.execute(
            """
            INSERT OR IGNORE INTO sample_population_tags(sample_id, tag, tag_order)
            VALUES (?, ?, ?)
            """,
            (sample_db_id, tag, tag_order),
        )

    return sample_db_id


def insert_mutation(connection, sample_db_id, mutation, reference):
    """Insert one mutation and its individual ALT/AF rows."""
    polymorphism = int(mutation.has_multiple_alt_alleles)
    repeats = strongest_alt_repeat_profile(mutation.alts)
    repeat_base, repeat_count = repeats[1]
    repeat_2_bases, repeat_2_count = repeats[2]
    repeat_3_bases, repeat_3_count = repeats[3]
    reference_6_before, reference_6_after = reference_context(
        reference,
        mutation.position,
    )
    metadata = {
        **mutation.metadata,
        "POLYMORPHISM": str(polymorphism),
        "REPEAT_1_BASE": repeat_base,
        "REPEAT_1_BASE_COUNT": str(repeat_count),
        "REPEAT_2_BASES": repeat_2_bases,
        "REPEAT_2_BASES_COUNT": str(repeat_2_count),
        "REPEAT_3_BASES": repeat_3_bases,
        "REPEAT_3_BASES_COUNT": str(repeat_3_count),
        "REFERENCE_6_BEFORE": reference_6_before,
        "REFERENCE_6_AFTER": reference_6_after,
    }
    cursor = connection.execute(
        """
        INSERT INTO mutations(
            sample_id, pos, ref, vcf_ref, alt, af, polymorphism, repeat_base,
            repeat_count, repeat_2_bases, repeat_2_count, repeat_3_bases,
            repeat_3_count, filter, metadata_json
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            sample_db_id,
            mutation.position,
            reference[mutation.position],
            mutation.ref,
            mutation.alt,
            mutation.allele_fraction_text,
            polymorphism,
            repeat_base,
            repeat_count,
            repeat_2_bases,
            repeat_2_count,
            repeat_3_bases,
            repeat_3_count,
            mutation.filter,
            json.dumps(metadata, sort_keys=True),
        ),
    )
    mutation_db_id = cursor.lastrowid

    for alt_index, alt in enumerate(mutation.alts):
        af_value = mutation.afs[alt_index] if alt_index < len(mutation.afs) else None
        numeric_af = af_value if isinstance(af_value, float) else None
        af_text = "" if af_value is None else str(af_value)
        alt_repeats = repeat_profile(alt)
        alt_repeat_base, alt_repeat_count = alt_repeats[1]
        alt_repeat_2_bases, alt_repeat_2_count = alt_repeats[2]
        alt_repeat_3_bases, alt_repeat_3_count = alt_repeats[3]
        connection.execute(
            """
            INSERT INTO mutation_alts(
                mutation_id, alt_index, alt, af, af_text, polymorphism,
                repeat_base, repeat_count, repeat_2_bases, repeat_2_count,
                repeat_3_bases, repeat_3_count
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                mutation_db_id,
                alt_index,
                alt,
                numeric_af,
                af_text,
                polymorphism,
                alt_repeat_base,
                alt_repeat_count,
                alt_repeat_2_bases,
                alt_repeat_2_count,
                alt_repeat_3_bases,
                alt_repeat_3_count,
            ),
        )


def make_sql_database(
    output_path,
    input_dir,
    reference_path=DEFAULT_MITO_REFERENCE_PATH,
    annotation_db_path=DEFAULT_ANNOTATION_DATABASE_PATH,
):
    """Create a SQLite database for subjects, samples, populations, and mutations."""
    samples = load_samples(input_dir)
    return make_sql_database_from_samples(
        output_path=output_path,
        samples=samples,
        reference_path=reference_path,
        annotation_db_path=annotation_db_path,
    )


def make_sql_database_from_plan(
    output_path,
    plan,
    reference_path=DEFAULT_MITO_REFERENCE_PATH,
    annotation_db_path=DEFAULT_ANNOTATION_DATABASE_PATH,
):
    """Create a SQLite database from editable import plan metadata."""
    samples = load_samples_from_plan(plan)
    return make_sql_database_from_samples(
        output_path=output_path,
        samples=samples,
        reference_path=reference_path,
        annotation_db_path=annotation_db_path,
    )


def make_sql_database_from_samples(
    output_path,
    samples,
    reference_path=DEFAULT_MITO_REFERENCE_PATH,
    annotation_db_path=DEFAULT_ANNOTATION_DATABASE_PATH,
):
    """Create a SQLite database from already-loaded samples."""
    reference = load_mito_reference(reference_path)
    output_path = Path(output_path)
    annotation_db_path = Path(annotation_db_path)
    if output_path.resolve() == annotation_db_path.resolve():
        raise ValueError(
            "The study output and persistent annotation database must use "
            "different paths."
        )

    with sqlite3.connect(output_path) as connection:
        create_database_schema(connection)
        skipped_out_of_reference = 0

        for sample in samples:
            subject_db_id = insert_subject(connection, sample.sample_id)
            sample_db_id = insert_sample(connection, sample, subject_db_id)
            for mutation in sample.mutations:
                if mutation.position not in reference:
                    skipped_out_of_reference += 1
                    continue
                insert_mutation(connection, sample_db_id, mutation, reference)

    annotate_imported_mutations(annotation_db_path, samples)

    if skipped_out_of_reference:
        print(
            f"Skipped {skipped_out_of_reference} mutation(s) outside the "
            f"{len(reference)} bp reference."
        )

    return output_path


def print_sample(sample, path):
    """Print a loaded sample and all of its parsed mutations."""
    total_variants, pass_variants, multi_alt_variants, _ = summarize_sample(sample)
    print(
        f"{path.name}: "
        f"{sample.label}: "
        f"{total_variants} variants "
        f"({pass_variants} PASS, {multi_alt_variants} multi-ALT)"
    )

    for mutation in sample.mutations:
        print(format_mutation(mutation))


def list_variant_files(input_dir):
    """Parse filtered VCFs and print each file's parsed mutations."""
    for path in iter_variant_files(input_dir):
        sample = load_sample(path, input_dir)
        print_sample(sample, path)


class ImportWindow:
    """Basic Tkinter importer for previewing files and editing tags."""

    def __init__(self, root):
        self.root = root
        self.root.title("Mito Variant Importer")
        self.plan = []
        self.items_by_iid = {}

        base_dir = Path(__file__).resolve().parent
        default_input = base_dir / "variant_calls" / "strict_filtered_calls"
        if not default_input.exists():
            default_input = base_dir

        self.input_dir = tk.StringVar(value=str(default_input))
        self.output_path = tk.StringVar(value=str(base_dir / "mito_import.sqlite"))
        self.annotation_db_path = tk.StringVar(
            value=str(DEFAULT_ANNOTATION_DATABASE_PATH)
        )
        self.reference_path = tk.StringVar(value=str(DEFAULT_MITO_REFERENCE_PATH))
        self.subject_regex = tk.StringVar(value=DEFAULT_SUBJECT_REGEX)
        self.subject_text = tk.StringVar()
        self.tags_text = tk.StringVar()
        self.annotation_progress = tk.DoubleVar(value=0)
        self.status_text = tk.StringVar(value="Load a preview, edit tags, then run the import.")

        self._build_widgets()
        self.load_preview()

    def _build_widgets(self):
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(1, weight=1)

        paths = ttk.Frame(self.root, padding=10)
        paths.grid(row=0, column=0, sticky="ew")
        paths.columnconfigure(1, weight=1)

        self._path_row(paths, 0, "Input", self.input_dir, self.browse_input_dir)
        self._path_row(paths, 1, "Output", self.output_path, self.browse_output_path)
        self._path_row(
            paths, 2, "Annotations", self.annotation_db_path,
            self.browse_annotation_db_path,
        )
        self._path_row(paths, 3, "Reference", self.reference_path, self.browse_reference_path)
        self._path_row(
            paths,
            4,
            "Subject regex",
            self.subject_regex,
            self.apply_subject_regex_to_all,
            button_text="Apply",
        )

        ttk.Button(paths, text="Refresh Preview", command=self.load_preview).grid(
            row=5,
            column=2,
            sticky="e",
            pady=(8, 0),
        )

        table_frame = ttk.Frame(self.root, padding=(10, 0, 10, 8))
        table_frame.grid(row=1, column=0, sticky="nsew")
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)

        columns = ("subject", "tags", "file", "status")
        self.tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show="headings",
            selectmode="extended",
        )
        self.tree.heading("subject", text="Subject")
        self.tree.heading("tags", text="Tags")
        self.tree.heading("file", text="File")
        self.tree.heading("status", text="Status")
        self.tree.column("subject", width=180, anchor="w")
        self.tree.column("tags", width=260, anchor="w")
        self.tree.column("file", width=420, anchor="w")
        self.tree.column("status", width=110, anchor="w")
        self.tree.grid(row=0, column=0, sticky="nsew")
        self.tree.bind("<<TreeviewSelect>>", self.sync_selection_fields)

        yscroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        yscroll.grid(row=0, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=yscroll.set)

        controls = ttk.Frame(self.root, padding=(10, 0, 10, 10))
        controls.grid(row=2, column=0, sticky="ew")
        controls.columnconfigure(1, weight=1)

        ttk.Label(controls, text="Subject ID").grid(row=0, column=0, sticky="w")
        ttk.Entry(controls, textvariable=self.subject_text).grid(
            row=0,
            column=1,
            sticky="ew",
            padx=6,
            pady=(0, 6),
        )
        ttk.Button(
            controls,
            text="Set Selected",
            command=self.replace_selected_subjects,
        ).grid(row=0, column=2, sticky="ew", padx=(0, 6), pady=(0, 6))

        ttk.Label(controls, text="Tags").grid(row=1, column=0, sticky="w")
        ttk.Entry(controls, textvariable=self.tags_text).grid(
            row=1,
            column=1,
            sticky="ew",
            padx=6,
        )
        ttk.Button(controls, text="Replace Selected", command=self.replace_selected_tags).grid(
            row=1,
            column=2,
            padx=(0, 6),
        )
        ttk.Button(controls, text="Add to Selected", command=self.add_selected_tags).grid(
            row=1,
            column=3,
        )

        footer = ttk.Frame(self.root, padding=(10, 0, 10, 10))
        footer.grid(row=3, column=0, sticky="ew")
        footer.columnconfigure(0, weight=1)

        ttk.Label(footer, textvariable=self.status_text).grid(row=0, column=0, sticky="w")
        self.go_button = ttk.Button(footer, text="Go", command=self.start_import)
        self.go_button.grid(row=0, column=1, padx=(8, 6))
        self.existing_button = ttk.Button(
            footer,
            text="Annotate Existing DBs",
            command=self.choose_existing_databases,
        )
        self.existing_button.grid(row=0, column=2, padx=(0, 6))
        self.clear_annotations_button = ttk.Button(
            footer,
            text="Clear Annotations",
            command=self.clear_annotations,
        )
        self.clear_annotations_button.grid(row=0, column=3, padx=(0, 6))
        ttk.Button(footer, text="Quit", command=self.root.destroy).grid(row=0, column=4)
        self.annotation_progress_bar = ttk.Progressbar(
            footer,
            variable=self.annotation_progress,
            maximum=1,
            mode="determinate",
        )
        self.annotation_progress_bar.grid(
            row=1,
            column=0,
            columnspan=5,
            sticky="ew",
            pady=(8, 0),
        )

    def _path_row(self, parent, row, label, variable, command, button_text="Browse"):
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", pady=2)
        ttk.Entry(parent, textvariable=variable).grid(
            row=row,
            column=1,
            sticky="ew",
            padx=6,
            pady=2,
        )
        ttk.Button(parent, text=button_text, command=command).grid(
            row=row,
            column=2,
            sticky="e",
            pady=2,
        )

    def browse_input_dir(self):
        path = filedialog.askdirectory(initialdir=self.input_dir.get() or ".")
        if path:
            self.input_dir.set(path)
            self.load_preview()

    def browse_output_path(self):
        path = filedialog.asksaveasfilename(
            initialfile=Path(self.output_path.get()).name,
            defaultextension=".sqlite",
            filetypes=[("SQLite databases", "*.sqlite *.sqlite3 *.db"), ("All files", "*.*")],
        )
        if path:
            self.output_path.set(path)

    def browse_reference_path(self):
        path = filedialog.askopenfilename(
            initialdir=str(Path(self.reference_path.get()).parent),
            filetypes=[("FASTA files", "*.fa *.fasta"), ("All files", "*.*")],
        )
        if path:
            self.reference_path.set(path)

    def browse_annotation_db_path(self):
        path = filedialog.asksaveasfilename(
            initialfile=Path(self.annotation_db_path.get()).name,
            defaultextension=".sqlite",
            filetypes=[("SQLite databases", "*.sqlite *.sqlite3 *.db"), ("All files", "*.*")],
        )
        if path:
            self.annotation_db_path.set(path)

    def derive_subject_id(self, item):
        return subject_id_from_regex(
            item.source_path,
            self.subject_regex.get() or DEFAULT_SUBJECT_REGEX,
        )

    def apply_subject_regex_to_item(self, iid):
        item = self.items_by_iid[iid]
        item.sample_id = self.derive_subject_id(item)
        self.tree.set(iid, "subject", item.sample_id)
        self.tree.set(iid, "status", "Ready")
        return item.sample_id

    def apply_subject_regex_to_all(self):
        try:
            re.compile(self.subject_regex.get() or DEFAULT_SUBJECT_REGEX)
        except re.error as exc:
            messagebox.showerror("Invalid subject regex", str(exc))
            return

        selected = self.tree.selection()
        target_iids = selected or self.tree.get_children()
        if not target_iids:
            messagebox.showinfo("No rows", "Load a preview before applying the regex.")
            return

        for iid in target_iids:
            self.apply_subject_regex_to_item(iid)

        self.sync_selection_fields()
        target = "selected" if selected else "previewed"
        self.status_text.set(
            f"Applied subject regex to {len(target_iids)} {target} file(s)."
        )

    def load_preview(self):
        try:
            self.plan = build_import_plan(Path(self.input_dir.get()))
            re.compile(self.subject_regex.get() or DEFAULT_SUBJECT_REGEX)
        except Exception as exc:
            messagebox.showerror("Preview failed", str(exc))
            self.status_text.set("Preview failed.")
            return

        self.tree.delete(*self.tree.get_children())
        self.items_by_iid.clear()
        for index, item in enumerate(self.plan):
            iid = str(index)
            self.items_by_iid[iid] = item
            item.sample_id = self.derive_subject_id(item)
            self.tree.insert(
                "",
                "end",
                iid=iid,
                values=(
                    item.sample_id,
                    ", ".join(item.population),
                    item.source_path.name,
                    "Ready",
                ),
            )

        self.status_text.set(f"Previewing {len(self.plan)} file(s).")

    def sync_selection_fields(self, _event=None):
        selected = self.tree.selection()
        if not selected:
            self.subject_text.set("")
            self.tags_text.set("")
            return

        if any(not self.items_by_iid[iid].sample_id for iid in selected):
            try:
                re.compile(self.subject_regex.get() or DEFAULT_SUBJECT_REGEX)
            except re.error as exc:
                messagebox.showerror("Invalid subject regex", str(exc))
                return

            for iid in selected:
                item = self.items_by_iid[iid]
                if not item.sample_id:
                    self.apply_subject_regex_to_item(iid)

        subjects = [
            self.items_by_iid[iid].sample_id
            for iid in selected
        ]
        first_subject = subjects[0]
        if all(subject == first_subject for subject in subjects):
            self.subject_text.set(first_subject)
        else:
            self.subject_text.set("")

        merged_tags = []
        for iid in selected:
            for tag in self.items_by_iid[iid].population:
                if tag not in merged_tags:
                    merged_tags.append(tag)

        self.tags_text.set(", ".join(merged_tags))
        if len(selected) > 1:
            self.status_text.set(
                f"Showing merged tags for {len(selected)} selected file(s)."
            )

    def replace_selected_subjects(self):
        subject_id = self.subject_text.get().strip()
        if not subject_id:
            messagebox.showinfo("Missing subject ID", "Enter a subject ID first.")
            return

        selected = self.tree.selection()
        if not selected:
            messagebox.showinfo("No rows selected", "Select one or more files first.")
            return

        for iid in selected:
            item = self.items_by_iid[iid]
            item.sample_id = subject_id
            self.tree.set(iid, "subject", item.sample_id)
            self.tree.set(iid, "status", "Ready")

        self.status_text.set(f"Updated subject ID for {len(selected)} file(s).")

    def replace_selected_tags(self):
        tags = split_tags(self.tags_text.get())
        self.update_selected_tags(lambda _existing: tags)

    def add_selected_tags(self):
        tags_to_add = split_tags(self.tags_text.get())

        def add_tags(existing):
            updated = list(existing)
            for tag in tags_to_add:
                if tag not in updated:
                    updated.append(tag)
            return updated

        self.update_selected_tags(add_tags)

    def update_selected_tags(self, update_func):
        selected = self.tree.selection()
        if not selected:
            messagebox.showinfo("No rows selected", "Select one or more files first.")
            return

        for iid in selected:
            item = self.items_by_iid[iid]
            item.population = update_func(item.population)
            self.tree.set(iid, "tags", ", ".join(item.population))
            self.tree.set(iid, "status", "Ready")

        self.status_text.set(f"Updated tags for {len(selected)} file(s).")

    def start_import(self):
        if not self.plan:
            messagebox.showinfo("Nothing to import", "Load a preview before importing.")
            return

        self.go_button.configure(state="disabled")
        self.existing_button.configure(state="disabled")
        self.clear_annotations_button.configure(state="disabled")
        for iid in self.items_by_iid:
            self.tree.set(iid, "status", "Queued")
        self.status_text.set("Import running...")

        thread = threading.Thread(target=self.run_import, daemon=True)
        thread.start()

    def choose_existing_databases(self):
        paths = filedialog.askopenfilenames(
            title="Choose previously imported study databases",
            initialdir=str(Path(self.output_path.get()).parent),
            filetypes=[
                ("SQLite databases", "*.sqlite *.sqlite3 *.db"),
                ("All files", "*.*"),
            ],
        )
        if not paths:
            return
        self.go_button.configure(state="disabled")
        self.existing_button.configure(state="disabled")
        self.clear_annotations_button.configure(state="disabled")
        self.annotation_progress.set(0)
        self.status_text.set(f"Annotating mutations from {len(paths)} database(s)...")
        threading.Thread(
            target=self.run_existing_annotation,
            args=(tuple(Path(path) for path in paths),),
            daemon=True,
        ).start()

    def run_existing_annotation(self, database_paths):
        try:
            annotation_path, allele_count = annotate_existing_databases(
                Path(self.annotation_db_path.get()),
                database_paths,
                progress_callback=self.report_annotation_progress,
            )
        except Exception as exc:
            self.root.after(0, self.existing_annotation_failed, str(exc))
            return
        self.root.after(
            0,
            self.existing_annotation_finished,
            annotation_path,
            allele_count,
            len(database_paths),
        )

    def report_annotation_progress(
        self, completed, total, position, ref, alt, provider, status
    ):
        self.root.after(
            0,
            self.update_annotation_progress,
            completed,
            total,
            position,
            ref,
            alt,
            provider,
            status,
        )

    def update_annotation_progress(
        self, completed, total, position, ref, alt, provider, status
    ):
        self.annotation_progress_bar.configure(maximum=max(total, 1))
        self.annotation_progress.set(completed)
        self.status_text.set(
            f"Annotations {completed}/{total}: "
            f"{position} {ref}>{alt}, {provider} ({status})"
        )

    def existing_annotation_finished(self, annotation_path, allele_count, db_count):
        self.go_button.configure(state="normal")
        self.existing_button.configure(state="normal")
        self.clear_annotations_button.configure(state="normal")
        message = (
            f"Processed {allele_count} unique allele(s) from {db_count} "
            f"database(s) into {annotation_path}"
        )
        self.status_text.set(message)
        messagebox.showinfo("Annotation complete", message)

    def existing_annotation_failed(self, error):
        self.go_button.configure(state="normal")
        self.existing_button.configure(state="normal")
        self.clear_annotations_button.configure(state="normal")
        self.status_text.set("Existing database annotation failed.")
        messagebox.showerror("Annotation failed", error)

    def clear_annotations(self):
        annotation_path = Path(self.annotation_db_path.get())
        confirmed = messagebox.askyesno(
            "Clear annotations?",
            "This will permanently remove every cached annotation from "
            f"{annotation_path}.\n\nContinue?",
        )
        if not confirmed:
            return
        try:
            variant_count, provider_count = clear_annotation_database(
                annotation_path
            )
        except Exception as exc:
            messagebox.showerror("Clear failed", str(exc))
            return
        self.annotation_progress.set(0)
        message = (
            f"Cleared {variant_count} variant(s) and {provider_count} "
            "provider annotation(s)."
        )
        self.status_text.set(message)
        messagebox.showinfo("Annotations cleared", message)

    def run_import(self):
        try:
            output_path = make_sql_database_from_plan(
                output_path=Path(self.output_path.get()),
                plan=self.plan,
                reference_path=Path(self.reference_path.get()),
                annotation_db_path=Path(self.annotation_db_path.get()),
            )
        except Exception as exc:
            self.root.after(0, self.import_failed, str(exc))
            return

        self.root.after(0, self.import_finished, output_path)

    def import_finished(self, output_path):
        for iid in self.items_by_iid:
            self.tree.set(iid, "status", "Imported")
        self.go_button.configure(state="normal")
        self.existing_button.configure(state="normal")
        self.clear_annotations_button.configure(state="normal")
        self.status_text.set(f"Wrote {output_path}")
        messagebox.showinfo("Import complete", f"Wrote {output_path}")

    def import_failed(self, error):
        for iid in self.items_by_iid:
            self.tree.set(iid, "status", "Error")
        self.go_button.configure(state="normal")
        self.existing_button.configure(state="normal")
        self.clear_annotations_button.configure(state="normal")
        self.status_text.set("Import failed.")
        messagebox.showerror("Import failed", error)


def launch_gui():
    global filedialog, messagebox, tk, ttk

    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk

    root = tk.Tk()
    ImportWindow(root)
    root.mainloop()


def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description="Build mutation SQLite/XLSX outputs from VCF files.",
    )
    parser.add_argument(
        "--gui",
        action="store_true",
        help="Open the interactive Tkinter importer.",
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        help="Directory containing VCF files for the study.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="SQLite output path.",
    )
    parser.add_argument(
        "--reference",
        type=Path,
        default=DEFAULT_MITO_REFERENCE_PATH,
        help=f"Mitochondrial reference FASTA. Default: {DEFAULT_MITO_REFERENCE_PATH}",
    )
    parser.add_argument(
        "--annotations-db",
        type=Path,
        default=DEFAULT_ANNOTATION_DATABASE_PATH,
        help=(
            "Persistent annotation cache shared by imports. "
            f"Default: {DEFAULT_ANNOTATION_DATABASE_PATH}"
        ),
    )
    parser.add_argument(
        "--annotate-existing",
        type=Path,
        nargs="+",
        metavar="DATABASE",
        help="Populate the annotation cache from existing imported databases.",
    )
    parser.add_argument(
        "--list",
        action="store_true",
        help="Print parsed VCF summaries instead of creating a SQLite database.",
    )
    return parser.parse_args(argv)

def file_in_folder_matching(path, match):
    """Return True if any parent folder name contains the given string."""
    dir_path = Path(path).resolve().parent
    while dir_path != dir_path.parent:
        if match in dir_path.name:
            return True
        dir_path = dir_path.parent
    return match in dir_path.name


if __name__ == "__main__":
    if len(sys.argv) == 1:
        launch_gui()
        raise SystemExit

    args = parse_args()
    if args.gui:
        launch_gui()
        raise SystemExit

    if args.annotate_existing:
        annotation_path, allele_count = annotate_existing_databases(
            args.annotations_db,
            args.annotate_existing,
        )
        print(
            f"Processed {allele_count} unique allele(s) into {annotation_path}"
        )
        raise SystemExit

    if args.input_dir is None:
        raise SystemExit("--input-dir is required outside of GUI mode.")

    if args.list:
        list_variant_files(input_dir=args.input_dir)
    else:
        if args.output is None:
            raise SystemExit("--output is required when creating a database.")

        output_path = make_sql_database(
            output_path=args.output,
            input_dir=args.input_dir,
            reference_path=args.reference,
            annotation_db_path=args.annotations_db,
        )
        print(f"Wrote {output_path}")
